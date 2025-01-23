import os
import glob
import shutil
import pyodbc
import openpyxl
import pandas as pd
from openpyxl import *
from xml.dom import minidom
from xml.dom.minidom import Node
from smbclient.path import (isdir)
from smbclient import open_file, walk
from datetime import timedelta, datetime
from openpyxl.styles import PatternFill, Font, Border, Side
'''************************************Основные вызываемые функции***********************************************'''
# Функция удаляет файлы в папке old и перемещает в нее вчерашний файл, на вход получает путь к папке Old
def delete_replace_files(folder_old: str, folder: str):
    # Удаление файла из папки Old
    try:
        os.remove(f'{folder_old}\\{os.listdir(folder_old)[0]}')  # Соединяется путь к папке и имя файла из каталога
    except IndexError:
        print('Ошибка при удалении файла, возможно файл не существует')
    # Перемещение вчерашнего файла в папку Old
    try:
        shutil.move(
            f'{folder}\\Jobs.xlsx',
            f'{folder_old}\\Jobs ({(datetime.now() - timedelta(days=1)).strftime("%d_%m_%Y")}).xlsx')
    except FileNotFoundError:
        print('Ошибка при перемещении файла. Файла не существует')
# Функция формирует таблицу и копирует ее в файл с макросами
def Create_table(File_xlsx: str, File_xlsm: str, File_routes: str):
    # Создается новый файл с джобами и в него записывается словарь
    with pd.ExcelWriter(File_xlsx) as writer:
        Parsing_data_to_dict().to_excel(writer, index=False, sheet_name='Джобы', engine='xlsxwriter')
    # Открываем исходный файл XLSX
    workbook_src = openpyxl.load_workbook(File_xlsx)
    # Открываем целевой файл XLSM
    workbook_dst = openpyxl.load_workbook(File_xlsm, keep_vba=True)
    # Объявление листов книги
    worksheet_1 = workbook_dst.worksheets[0]
    worksheet_2 = workbook_dst.worksheets[1]
    # Очистка листов от всех записей
    worksheet_1.delete_rows(5, worksheet_1.max_row)
    worksheet_2.delete_rows(5, worksheet_2.max_row)
    # Копируем данные из исходного листа в целевой лист, вставляя с 4 строки
    for row in range(1, workbook_src.active.max_row + 1):
        for col in range(1, workbook_src.active.max_column + 1):
            worksheet_1.cell(row=row + 3, column=col, value=workbook_src.active.cell(row=row, column=col).value)
    # Вызов функции для создания файла с проходами джобов
    SQL_create_table(File_routes)
    # Открываем лист таблицы с проходами джобов, который создает функция SQL_create_table
    wsJobConnect = openpyxl.load_workbook(File_routes)
    # Вставляем во 2 лист таблицы данные о проходах джобов с 5 строки
    for row in range(1, wsJobConnect.active.max_row + 1):
        for col in range(1, wsJobConnect.active.max_column + 1):
            worksheet_2.cell(row=row + 4, column=col, value=wsJobConnect.active.cell(row=row, column=col).value)
    # Сохраняем целевой файл XLSM
    workbook_dst.save(File_xlsm)
    print('Заполнена таблица с макросами Jobs (VBA).xlsm')
# Функция форматирует таблицу с макросами
def Custom_table(Filename: str):
    # openpyxl открывает файл, инициируется переменная с листами таблицы
    workbook = load_workbook(Filename, keep_vba=True)
    worksheet = workbook['Джобы']
    worksheet2 = workbook['Проходы в системы']
    # Добавляется фильтр к полям двух листов
    worksheet.auto_filter.ref = f'A4:J{worksheet.max_row}'
    worksheet2.auto_filter.ref = f'A4:H{worksheet.max_row}'
    # Задается ширина полей первого листа. (значение в символах)
    worksheet.column_dimensions['A'].width = 45
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['C'].width = 13
    worksheet.column_dimensions['D'].width = 50
    worksheet.column_dimensions['E'].width = 30
    worksheet.column_dimensions['F'].width = 10
    worksheet.column_dimensions['G'].width = 10
    worksheet.column_dimensions['H'].width = 10
    worksheet.column_dimensions['I'].width = 25
    worksheet.column_dimensions['J'].width = 25
    # Задается ширина полей второго листа. (значение в символах)
    worksheet2.column_dimensions['A'].width = 45
    worksheet2.column_dimensions['B'].width = 40
    worksheet2.column_dimensions['C'].width = 30
    worksheet2.column_dimensions['D'].width = 15
    worksheet2.column_dimensions['E'].width = 25
    worksheet2.column_dimensions['F'].width = 6
    worksheet2.column_dimensions['G'].width = 25
    worksheet2.column_dimensions['H'].width = 39
    #  Цикл по каждой строке первого листа, кроме строки с именами полей
    for row in range(5, worksheet.max_row + 1):
        # Меняется размер шрифта для каждого столбца
        worksheet[f'A{row}'].font = Font(size=9)
        worksheet[f'B{row}'].font = Font(size=9)
        worksheet[f'C{row}'].font = Font(size=9)
        worksheet[f'D{row}'].font = Font(size=9)
        worksheet[f'E{row}'].font = Font(size=9)
        worksheet[f'F{row}'].font = Font(size=9)
        worksheet[f'G{row}'].font = Font(size=9)
        worksheet[f'H{row}'].font = Font(size=9)
        worksheet[f'I{row}'].font = Font(size=9)
        worksheet[f'J{row}'].font = Font(size=9)
        # Границы ячеек
        worksheet[f'A{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'B{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'C{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'D{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'E{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'F{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'G{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'H{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'I{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet[f'J{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                             top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        # Меняется заливка для каждого столбца
        worksheet[f'A{row}'].fill = PatternFill('solid', fgColor='00B050')
        worksheet[f'B{row}'].fill = PatternFill('solid', fgColor='DCE6F1')
        worksheet[f'C{row}'].fill = PatternFill('solid', fgColor='FABF8F')
        worksheet[f'D{row}'].fill = PatternFill('solid', fgColor='00B0F0')
        worksheet[f'E{row}'].fill = PatternFill('solid', fgColor='E4DFEC')
        worksheet[f'F{row}'].fill = PatternFill('solid', fgColor='CCC0DA')
        worksheet[f'J{row}'].fill = PatternFill('solid', fgColor='D9D9D9')
        # Добавление гиперссылок к логам джобов
        worksheet[f'D{row}'].hyperlink = worksheet[f'D{row}'].value
    #  Цикл по каждой строке второго листа, кроме строки с именами полей
    for row in range(5, worksheet2.max_row + 1):
        # Меняется размер шрифта для каждого столбца
        worksheet[f'A{row}'].font = Font(size=9)
        worksheet[f'B{row}'].font = Font(size=9)
        worksheet[f'C{row}'].font = Font(size=9)
        worksheet[f'D{row}'].font = Font(size=9)
        worksheet[f'E{row}'].font = Font(size=9)
        worksheet[f'F{row}'].font = Font(size=9)
        worksheet[f'G{row}'].font = Font(size=9)
        worksheet[f'H{row}'].font = Font(size=9)
        # Границы ячеек
        worksheet2[f'A{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'B{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'C{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'D{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'E{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'F{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'G{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        worksheet2[f'H{row}'].border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        # Меняется заливка для каждого столбца
        worksheet2[f'A{row}'].fill = PatternFill('solid', fgColor='00B050')
        worksheet2[f'B{row}'].fill = PatternFill('solid', fgColor='DCE6F1')
        worksheet2[f'C{row}'].fill = PatternFill('solid', fgColor='FABF8F')
        worksheet2[f'D{row}'].fill = PatternFill('solid', fgColor='00B0F0')
        worksheet2[f'E{row}'].fill = PatternFill('solid', fgColor='E4DFEC')
        worksheet2[f'F{row}'].fill = PatternFill('solid', fgColor='CCC0DA')
        worksheet2[f'G{row}'].fill = PatternFill('solid', fgColor='D9D9D9')
        worksheet2[f'H{row}'].fill = PatternFill('solid', fgColor='FFE699')
    # Сохранение таблицы
    workbook.save(Filename)
    print('Отформатирована таблица с макросами Jobs(VBA).xlsm')
'''******************************************************************************************************************'''
'''---------------------------------Вспомогательные глобальные переменные и функции----------------------------------'''
# Глобальный список с именами джобов
jobs_name = []
# Функция парсит данные с серверов и сохраняет их в словарь для дальнейшего добавления в таблицу
def Parsing_data_to_dict():
    # Пустой фрейм для записи в таблицу
    data_table = pd.DataFrame({
        'JobName': [],
        'Описание': [],
        'Server': [],
        'LogFile': [],
        'TaskName': [],
        'StartType': [],
        'Start': [],
        'Stop': [],
        'bat file': [],
        'Repository Path': []
    })
    # Цикл проходит по всем серверам с помощью функции парсера и добавляет данные о джобах в пустой фрейм data_table
    for server in ['server', 'server1', 'server3', 'server4']:
        data_table = data_table.append(parseJobs(server, f'N:\\folder\\path\\Jobs\\Files\\file.xlsx'), ignore_index=True)
    print('Сформирована таблица с джобами Jobs.xlsx')
    return data_table
# Функция парсит данные о джобах с сервера в промежуточный фрейм
def parseJobs(server: str, Info_file: str):
    # Пустой фрейм, для записи данных о джобах с сервера
    result = pd.DataFrame({
        'JobName': [],
        'Описание': [],
        'Server': [],
        'LogFile': [],
        'TaskName': [],
        'StartType': [],
        'Start': [],
        'Stop': [],
        'bat file': [],
        'Repository Path': []
    })
    # Определение пути к папке с джобами на сервере. На сервере server7 путь отличается
    if server == "server7":
        fullpath = f"\\\\{server}\\c$\\Windows\\System32\\Tasks\\\Microsoft\\START JOBS"
    else:
        fullpath = f"\\\\{server}\\c$\\Windows\\System32\\Tasks\\START JOBS"
    for file in walk(fullpath):
        for filename in file[2]:  # Цикл для обхода всех объектов в папке. Текущий объект помещается в filename.
            fullpathfile = f"{file[0]}\\{filename}"  # Формируем строку с полным именем до объекта (файла).
            if not isdir(fullpathfile):  # Проверяем не является ли объект папкой и только тогда начинаем обработку.
                enableCheck = 'true'  # Переменная для проверки сосояния джоба
                dom = minidom.parse(fullpathfile)  # Парсим  XML файл джоба.
                try:
                    enableCheck = dom.getElementsByTagName('Settings')[0].getElementsByTagName("Enabled")[
                        0].firstChild.data  # Проверяем наличие тэга отключенности джоба.
                except IndexError:
                    pass
                if enableCheck == "true":  # Если джоб включен, то парсим данные дальше
                    # Инициализация переменных
                    allwaysupName, startType, resultJobName, resultLogFile, resultRepPath, schedule = (
                        "NULL", "NULL", "NULL", "NULL", "NULL", "NULL")
                    # Парсим имя таски
                    if 'Taskkill SPOON' in dom.getElementsByTagName('URI')[0].firstChild.data.split('\\'):
                        continue
                    else:
                        TaskName = dom.getElementsByTagName('URI')[0].firstChild.data.split('\\')[-1]
                    # Получаем путь к исполняемому файлу из таски
                    execComand = dom.getElementsByTagName('Command')[0].firstChild.data.strip('"')
                    # Собираем сведения о расписании (Запуск):
                    # Инициализация переменных
                    schedule, time, interval, duration, days = "", "", "", "", ""
                    # Парсим тэг, содержащий информацию о времени запуска
                    triggers = dom.getElementsByTagName('CalendarTrigger')
                    for elem1 in triggers:
                        # время запуска джоба
                        time = elem1.getElementsByTagName('StartBoundary')[0].firstChild.data.split('T')[-1]
                        # интервал запуска джоба
                        schedule = schedule + time + " "
                        # Сведения о запуске по дням недели. Ищем тэг DaysOfWeek
                        for elem in elem1.getElementsByTagName('DaysOfWeek'):
                            for y in elem.childNodes:
                                if y.nodeType == Node.ELEMENT_NODE:
                                    day = str(y)
                                    schedule = schedule + day[14:day.rfind("at")]
                    # Собираем сведения о расписании (Остановка). Так как файлы с описанием остановки заданий находятся в другой папке,
                    # выполняется поиск соответствующего файла и парсится аналогичным образом как файл запуска:
                    # Инициализация переменных
                    scheduleStop, timeStop, intervalStop, durationStop, daysStop = "", "", "", "", ""
                    # Путь до папки с отключением джобов
                    fullpathStop = fullpath.replace("\START", "\STOP")
                    # Путь до джобов STOP
                    fullpathStopFile = f"{fullpathStop}\\*{filename[10:]}"
                    # Список с путями файлов, находящихся в папке с джобами
                    findStopFile = glob.glob(fullpathStopFile)
                    if not findStopFile:
                        timeStop = "Не найдено"
                    else:
                        # Парсим XML файл джоба
                        domStop = minidom.parse(str(findStopFile[0]))
                        # Парсим тэг CalendarTrigger
                        triggersStop = domStop.getElementsByTagName('CalendarTrigger')
                        for elem1 in triggersStop:
                            # Время остановки джоба
                            timeStop = elem1.getElementsByTagName('StartBoundary')[0].firstChild.data.split('T')[-1]
                    if '.bat' in execComand:  # Проверяется если файл это батник
                        # Если в пути к файлу нет '.bat' или 'C:\Pentaho' идем дальше
                        if '.bat' not in execComand or 'C:\Pentaho' not in execComand:
                            continue
                        # Получаем путь до батника с джобом
                        shareBatPath = f"\\\\{server}\\c$\\{execComand[3:]}"
                        startType = "Scheduler"  # Тип запуска джоба
                    elif 'cmd' in execComand:  # Если в задании путь до .cmd, начинаем получение пути до cmd файла
                        # Если в пути cmd файла нет Control или 'C:\Pentaho', этот cmd не от Pentaho джоба
                        if "\Control" not in execComand or 'C:\Pentaho' not in execComand:
                            continue
                        startType = "AlwaysUp"  # Тип запуска джоба
                        # Открываем cmd файл, проходим по каждой строке и парсим имя джоба:
                        with open_file(f"\\\\{server}\\c$\\{execComand[3:]}") as cmdFile:
                            for line in cmdFile:
                                if 'net start' in line:
                                    TaskName = line.split('\"')[-2]
                        # Используем файл соответствия сервисов и батников для определения нужного батника.
                        shareBatPath = getBatFromService(server, TaskName)
                        if shareBatPath == "Disabled" or shareBatPath == "Error":
                            continue
                        # Присваивается переменной имя таски из AlwaysUp
                        TaskName = TaskName.split(' (')[0]
                    # Парсится батник. Пропускаются пустые строки. Ищется строка без комментариев.
                    with open_file(shareBatPath) as batFile:
                        for line in batFile:  # Перебираются все строки в батнике
                            if line == '\n' or line == '' or line == ' ' or line == '   ' or line[0] == ':' or line[
                                1] == 'd': continue
                            words = line.split(' ')  # Разбили строку на список с разделителем - пробел.
                            if 'data-integration' in words[0]:  # Запуск происходит напрямую
                                jobName = words[2].split(':')[1].strip('\"')
                            if 'runjob' in words[0]:  # Запуск происходит через runjob
                                jobName = words[2]
                            # Имя джоба
                            resultJobName = jobName
                            # Добавление имени джоба в глобальную переменную jobs_name
                            jobs_name.append(resultJobName)
                            for log in words:  # Ищем файл лога
                                logPath = 'NULL'
                                if ".log" in log:  # Ищим путь к логу и подставляем путь к серверу
                                    logPath = f"\\\\{server}{log[11:]}".replace("\n", "")
                                    break
                            resultLogFile = logPath
                            for rep in words:  # Ищем путь в репозитории
                                repPath = 'NULL'
                                if "Production/" in rep or "Dev/" in rep:
                                    repPath = rep
                                    break
                            resultRepPath = repPath.replace("/dir:", "")
                            resultRepPath = resultRepPath.replace('"', '').replace('/Production', 'Production')
                    '''----------------------------------------------------------------------------------------------'''
                    jobs_description = {}  # Инициализация словаря с описанием джобов из отдельного файла
                    # Открытие листа из книги с описанием джобов
                    jobs_info = openpyxl.load_workbook(Info_file)
                    worksheet = jobs_info.active
                    # Цикл по таблице, где в словарь jobs_description добавляются значения из 2 поля с ключем из 1 поля
                    for i in worksheet.values:
                        jobs_description[i[0]] = i[1]
                    info = ''
                    # Если название джоба есть в ключах словаря с описанием, то info присваивается описание джоба
                    if resultJobName in jobs_description.keys():
                        info = jobs_description[resultJobName]
                    '''----------------------------------------------------------------------------------------------'''
                    new_row = {
                        "JobName": resultJobName,
                        "Описание": info,
                        "Server": server,
                        "LogFile": resultLogFile,
                        'TaskName': TaskName,
                        "StartType": startType,
                        "Start": schedule,
                        "Stop": timeStop,
                        "bat file": shareBatPath.replace(f"\\\\{server}\\c$", "C:"),
                        "Repository Path": resultRepPath,
                    }
                    # Добавление словаря с информацие о джобах в итоговый словарь о сервере
                    result = result.append(new_row, ignore_index=True)
                    print(f'********COMPLETED SUCCESSFULLY********\n"JobName": {resultJobName}\n"Описание":{info}\n'
                          f'"Server": {server}\n"LogFile": {resultLogFile}\n"TaskName": {TaskName}\n'
                          f'"StartType": {startType}\n"Start": {schedule}\n"Stop": {timeStop}\n'
                          f'"bat file": {shareBatPath}\n"Repository Path": {resultRepPath}\n')
    return result
# Функция находит путь к исполняемому файлу джоба на сервере, принимает путь к папке с джобами и имя джоба
def getBatFromService(server: str, serviceName: str):
    # Открытие CSV файла соответствия сервисов с сервера
    csv = pd.read_csv(f"\\\\{server}\\D$\\Services\\Services.csv", sep=',', skiprows=1, index_col=0)
    try:
        PathName = csv.loc[[serviceName], 'PathName']
    except Exception:
        return "Error"
    check = csv.loc[[serviceName], 'StartMode'][serviceName]
    if check != "Disabled":
        command = PathName[serviceName]
        if 'Pentaho9' in PathName[0]:
            batPath = command[command.find("C:\\Pentaho\\"):command.find(".bat") + 4]
        else:
            batPath = command[command.find("C:\\Pentaho6\\"):command.find(".bat") + 4]
        fullBatPath = f"\\\\{server}\\c$\\{batPath[3:]}"
        return fullBatPath
    else:
        return "Disabled"
# Функция посылает запрос к БД и формирует таблицу с информацией о проходах к джобам
def SQL_create_table(Routes_file: str):
    # Удаляем прошлый файл
    try:
        os.remove(Routes_file)
    except:
        pass
    # Создаем книгу
    wb = Workbook()
    # Открываем активный лист
    sh = wb.active
    # Формируем подключение к БД
    connection = pyodbc.connect(r'Driver={SQL Server};Server=Server;Database=name;Trusted_Connection=yes')
    # Создаем курсор
    cursor = connection.cursor()
    # Направляем запрос к БД
    cursor.execute("""select DISTINCT
                        t.NAME, j.NAME as JOB, CAST(ja_d.VALUE_STR AS VARCHAR(255)) as DIR_PATH, d.DIRECTORY_NAME
                        , db.ID_DATABASE, db.NAME, db.ID_TYPE, db.ID_CONTYPE, db.HOST_NAME, CAST(db.DATABASE_NAME AS VARCHAR(255)) DATABASE_NAME, db.PORT, db.USERNAME, db.PASSWORD, db.SERVERNAME, db.DATA_TBS, db.INDEX
                        --, d2.*
                        from R_T t (nolock)
                        join R_JOBEN ja (nolock) on (ja.CODE = 'name' and cast(ja.VALUE_STR as varchar(50)) = t.NAME)
                        join R_JOBEB_ATT ja_d (nolock) on (ja.ID_JOBEN = ja_d.ID_JOBEN and ja_d.CODE = 'dir_path')
                        join R_JOB j (nolock) on (j.ID_JOB = ja.ID_JOB)
                        join R_DIRECTORY d (nolock) on (d.ID_DIRECTORY = j.ID_DIRECTORY)
                        join R_DATABASE st (nolock) on (st.ID_TRANSF = t.ID_TRANSF)
                        join R_DATABASE db (nolock) on (db.ID_DATABASE = st.ID_DATABASE)
                        JOIN R_DIRECTORY d2 (nolock) ON t.ID_DIRECTORY = d.ID_DIRECTORY
                        --WHERE j.NAME = 'Job_Load'
                        --AND d.DIRECTORY_NAME = 'CB_R'
                        --ORDER BY 1, 4
                        UNION ALL
                        SELECT T3.NAME, T1.NAME, NULL, NULL, T5.*
                        from R_JOB as T1
                        JOIN R_JOB_HOP as T2 ON T1.ID_JOB = T2.ID_JOB
                        JOIN R_JOB as T3 ON T2.ID_JOB_COPY_TO = T3.ID_JOB
                        JOIN R_JOB_DATABASE as T4 ON T3.ID_JOB = T4.ID_JOB
                        JOIN R_DATABASE as T5 ON T4.ID_DATABASE = T5.ID_DATABASE
                        ORDER BY 2, 1""")
    # Получаем ответ от БД в виде списка кортежей
    rows = cursor.fetchall()
    # Сбрасываем запрос
    cursor.close()
    # Создаем курсор
    cursor = connection.cursor()
    # Отправляем новый запрос на проверку смотрит ли джоб в другой джоб
    cursor.execute("""SELECT T1.NAME as Job_name, T3.NAME as Job_transform
                            from R_JOB as T1
                            JOIN R_JOB_HOP as T2 ON T1.ID_JOB = T2.ID_JOB
                            JOIN R_JOBEN as T3 ON T2.ID_JOBEN_COPY = T3.ID_JOBEN
                            WHERE T3.NAME in (select NAME from R_JOB)""")
    # Сохраняем эти данные в список в виде (Название джоба, название джоба в который смотрит)
    JobsInJobs = cursor.fetchall()
    # Список для проверки уникальных строк
    data_unique = []
    # Цикл по строкам из запроса
    for row in rows:
        # Если строка есть в списке для проверки, то идем дальше
        if row in data_unique:
            continue
        # Если в строке нет слова test или old, строки нет в списке уникальности и имя джоба есть в глобальной переменной, то добавляем строку в таблицу
        if "test" not in ''.join(str(i) for i in row).lower() and "old" not in ''.join(
                str(i) for i in row).lower() and row[1].lower().strip() in map(lambda x: x.lower().strip(), jobs_name):
            sh.append((row[1], row[0], row[5], row[8], row[9], row[10], row[11]))
            # Добавляем строку в список для проверки уникальности
            data_unique.append(row)
        # Вложенный цикл по джобам к которым обращаются другие джобы
        for job in JobsInJobs:
            if (job[1].lower().strip() == row[1].lower().strip() and "test" not in ''.join(
                    str(i) for i in row).lower() and "old" not in ''.join(str(i) for i in row).lower()
                    and row not in data_unique and 'old' not in ''.join(str(i) for i in job)):
                sh.append((job[0], row[0], row[5], row[8], row[9], row[10], row[11], job[1]))
                data_unique.append(row)
    # Сохраняем таблицу
    wb.save(Routes_file)
    print('Сформирована таблица с проходами Jobs routes.xlsx')
    # Закрываем соединение с БД
    connection.close()
'''------------------------------------------------------------------------------------------------------------------'''
